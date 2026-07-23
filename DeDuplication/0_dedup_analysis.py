# DeDuplication/0_dedup_analysis.py
# Run with: python -m DeDuplication.0_dedup_analysis

import hashlib
import logging
import re
import sys
from datetime import datetime
from pathlib import Path

import ollama
import pandas as pd

# Add project root to path so we can import the central config
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from project_config import (
    SOURCE_DOCS_DIR as SOURCE_DIR,      # ← Single source of truth
    DEDUPS_DIR,
    TRIVIAL_SUBJECTS as TRIVIAL_SUBJECTS_FILE,
)

# Optional libraries for broader extraction
try:
    import fitz  # PyMuPDF for PDFs
except ImportError:
    fitz = None
try:
    from docx import Document
except ImportError:
    Document = None

from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np
from openpyxl.styles import PatternFill

# ========================= CONFIG =========================
SIMILARITY_THRESHOLD = 0.95

timestamp = datetime.now().strftime("%Y%m%d_%H%M")
OUTPUT_EXCEL = DEDUPS_DIR / f"deduplication_review_{timestamp}.xlsx"
LOG_FILE = DEDUPS_DIR / f"dedup_analysis_{timestamp}.log"

# ========================= LOGGING =========================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

logger.info("=== Deduplication Analysis Started ===")
logger.info(f"Source directory: {SOURCE_DIR}")
logger.info(f"Output folder: {DEDUPS_DIR}")
logger.info(f"Trivial subjects file: {TRIVIAL_SUBJECTS_FILE}")

# Load trivial subjects
if TRIVIAL_SUBJECTS_FILE.exists():
    trivial_subjects = [line.strip() for line in TRIVIAL_SUBJECTS_FILE.read_text(encoding="utf-8").splitlines() if line.strip()]
    logger.info(f"Loaded {len(trivial_subjects)} trivial subjects")
else:
    trivial_subjects = []
    logger.warning("trivial_subjects.txt not found – trivial detection disabled")

# Load embedder
logger.info("Loading embedding model...")
embedder = SentenceTransformer('Lajavaness/bilingual-embedding-small', trust_remote_code=True)


def extract_text_from_file(file_path: Path) -> str:
    suffix = file_path.suffix.lower()
    try:
        if suffix == ".txt":
            return file_path.read_text(encoding="utf-8", errors="ignore")
        elif suffix == ".pdf" and fitz:
            doc = fitz.open(file_path)
            text = "\n".join(page.get_text("text") for page in doc)
            doc.close()
            return text
        elif suffix in (".docx", ".doc") and Document:
            doc = Document(file_path)
            return "\n".join(p.text for p in doc.paragraphs)
        else:
            return file_path.name
    except Exception as e:
        logger.warning(f"Text extraction failed for {file_path.name}: {e}")
        return file_path.name


def extract_hyperlinks(text: str) -> str:
    pattern = r'(https?://\S+|www\.\S+|\b[a-zA-Z0-9-]+\.(com|ca|org|net|gov|edu)\b)'
    links = re.findall(pattern, text, re.IGNORECASE)
    cleaned = set(link[0] if isinstance(link, tuple) else link for link in links)
    return ", ".join(sorted(cleaned)) if cleaned else ""


def is_trivial_content(text: str) -> bool:
    if not trivial_subjects or not text.strip():
        return False
    prompt = f"""Determine if the following document content is trivial/non-business related. 
It is trivial if it matches ANY of these topics: {', '.join(trivial_subjects[:40])} ... (and the rest).

Document text (first 2000 characters):
{text[:2000]}

Answer only with Yes or No."""
    try:
        response = ollama.chat(model="qwen2.5:7b", messages=[{"role": "user", "content": prompt}])
        answer = response["message"]["content"].strip().lower()
        return "yes" in answer
    except Exception as e:
        logger.warning(f"Trivial content detection failed: {e}")
        return False


def compute_file_hash(file_path: Path) -> str:
    hash_sha = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            hash_sha.update(chunk)
    return hash_sha.hexdigest()


# ========================= MAIN ANALYSIS =========================
logger.info(f"Scanning source directory: {SOURCE_DIR}")
files = list(SOURCE_DIR.rglob("*.*"))
logger.info(f"Found {len(files)} files to analyze")

records = []
embeddings = []
file_paths = []

for file_path in files:
    if file_path.is_file():
        try:
            stat = file_path.stat()
            hash_value = compute_file_hash(file_path)
            text = extract_text_from_file(file_path)
            is_trivial = is_trivial_content(text)

            emb = embedder.encode(text[:3000], normalize_embeddings=True)

            records.append({
                "filename": file_path.name,
                "original_path": str(file_path),
                "size_mb": round(stat.st_size / (1024 * 1024), 2),
                "creation_time": datetime.fromtimestamp(stat.st_ctime).strftime("%Y-%m-%d %H:%M"),
                "last_modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                "Is_Trivial": "Yes" if is_trivial else "No",
                "Hash_Value": hash_value,
            })
            embeddings.append(emb)
            file_paths.append(file_path)
        except Exception as e:
            logger.warning(f"Failed to process {file_path.name}: {e}")

if len(embeddings) == 0:
    logger.error("No files processed.")
    print("No files found to analyze.")
    exit()

df = pd.DataFrame(records)
embeddings = np.array(embeddings)

logger.info("Computing similarity matrix...")
sim_matrix = cosine_similarity(embeddings)

# ==================== IMPROVED CLUSTERING LOGIC ====================
visited = set()
cluster_id = 0
df["Cluster_ID"] = "Unique"
df["Is_Master"] = False
df["Similarity_Score"] = 1.0
df["Similarity_Found"] = ""
df["Discrepancy"] = ""

for i in range(len(df)):
    if i in visited:
        continue

    cluster = [i]
    visited.add(i)

    for j in range(i + 1, len(df)):
        if sim_matrix[i, j] >= SIMILARITY_THRESHOLD and j not in visited:
            cluster.append(j)
            visited.add(j)

    if len(cluster) > 1:
        # Real duplicate cluster
        cluster_id += 1
        cid = f"DUP-{cluster_id:04d}"

        cluster_files = [(idx, file_paths[idx]) for idx in cluster]
        # Oldest file becomes master
        master_idx = min(cluster_files, key=lambda x: (
            Path(x[1]).stat().st_ctime,
            Path(x[1]).stat().st_mtime
        ))[0]

        for idx in cluster:
            df.loc[idx, "Cluster_ID"] = cid
            df.loc[idx, "Is_Master"] = (idx == master_idx)
            score = float(sim_matrix[master_idx, idx])
            df.loc[idx, "Similarity_Score"] = round(score, 4)

            if abs(score - 1.0) < 0.0001 and df.loc[idx, "Hash_Value"] == df.loc[master_idx, "Hash_Value"]:
                df.loc[idx, "Similarity_Found"] = "Exact duplicate (hash + content)"
            else:
                df.loc[idx, "Similarity_Found"] = f"{score:.2%} semantic similarity"

            # Discrepancy only for duplicates
            master_path = Path(df.loc[master_idx, "original_path"])
            current_path = Path(df.loc[idx, "original_path"])
            if df.loc[idx, "filename"] != df.loc[master_idx, "filename"]:
                df.loc[idx, "Discrepancy"] = "Different filename"
            elif master_path.stat().st_mtime != current_path.stat().st_mtime:
                df.loc[idx, "Discrepancy"] = "Different modification date"
            else:
                df.loc[idx, "Discrepancy"] = "Minor content variations"
    else:
        # True unique file
        df.loc[i, "Similarity_Found"] = "Unique file"
        df.loc[i, "Discrepancy"] = "No similar documents found"
        df.loc[i, "Is_Master"] = True

# ====================== RECOMMENDED ACTION ======================
def get_recommended_action(row):
    if row["Is_Trivial"] == "Yes":
        return "Review (trivial content)"
    if row["Cluster_ID"] == "Unique" or row["Is_Master"]:
        return "Keep as Master"
    return "Delete"

df["Recommended_Action"] = df.apply(get_recommended_action, axis=1)
df["User_Confirmed_Delete"] = df.apply(
    lambda row: "Yes" if row["Recommended_Action"] == "Delete" else "", axis=1
)

df["Exact_Duplicate"] = df.duplicated(subset=["Hash_Value"], keep=False)
df["Near_Duplicate"] = (df["Similarity_Score"] >= SIMILARITY_THRESHOLD) & (df["Cluster_ID"] != "Unique")

# Add Hash_Type column (was missing)
df["Hash_Type"] = "SHA-256"

# Add downstream hyperlinks
df["downstream_hyperlinks"] = df["original_path"].apply(
    lambda p: extract_hyperlinks(extract_text_from_file(Path(p)))
)

# Final column order (now includes Hash_Type)
cols = [
    "Cluster_ID", "Is_Master", "filename", "original_path", "size_mb",
    "creation_time", "last_modified", "Exact_Duplicate", "Near_Duplicate",
    "Similarity_Score", "Similarity_Found", "Discrepancy",
    "Recommended_Action", "User_Confirmed_Delete", "Is_Trivial",
    "downstream_hyperlinks", "Hash_Value", "Hash_Type"
]

df = df[cols]

# ========================= SAVE TO EXCEL =========================
with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Duplicate_Clusters", index=False)

    # Hyperlink filename column + color Recommended_Action
    ws = writer.sheets["Duplicate_Clusters"]
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for row in range(2, len(df) + 2):
        # Filename hyperlink
        cell = ws.cell(row=row, column=3)  # filename is column C (index 3)
        link = df.iloc[row - 2]["original_path"]
        cell.hyperlink = link
        cell.value = df.iloc[row - 2]["filename"]
        cell.style = "Hyperlink"

        # Color Recommended_Action column (column M = index 13)
        action_cell = ws.cell(row=row, column=13)
        if action_cell.value in ("Delete", "Review (trivial content)"):
            action_cell.fill = red_fill

    # Summary sheet
    summary = pd.DataFrame({
        "Metric": [
            "Total Files Scanned",
            "Duplicate Clusters Found",
            "Files Recommended for Deletion",
            "Trivial Content Flagged",
            "Unique Files",
            "Run Timestamp"
        ],
        "Value": [
            len(df),
            df["Cluster_ID"].nunique() - (1 if "Unique" in df["Cluster_ID"].values else 0),
            (df["Recommended_Action"] == "Delete").sum(),
            (df["Is_Trivial"] == "Yes").sum(),
            (df["Cluster_ID"] == "Unique").sum(),
            datetime.now().strftime("%Y-%m-%d %H:%M")
        ]
    })
    summary.to_excel(writer, sheet_name="Summary", index=False)

logger.info(f"Analysis complete. Excel report saved to: {OUTPUT_EXCEL}")
print(f"\n✅ Deduplication analysis complete!")
print(f"   Excel:  {OUTPUT_EXCEL}")
print(f"   Log:    {LOG_FILE}")
print(f"   Source: {SOURCE_DIR}")