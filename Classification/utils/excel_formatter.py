# Classification/utils/excel_formatter.py
"""
Applies formatting to the classification_results.xlsx output:
- Freeze header row
- Enable auto-filter
- Color rows based on confidence_category (High/Medium/Low)
"""

import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)

# Confidence-based row colors (light shades for readability)
FILL_COLORS = {
    "High":   PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Medium": PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid"),
    "Low":    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}


def format_classification_excel(excel_path: Path) -> None:
    """
    Apply standard formatting to the classification Excel file.
    - Freeze first row
    - Add auto-filter
    - Color rows based on 'confidence_category' column

    Does nothing if file is missing or cannot be processed.
    """
    if not excel_path.is_file():
        logger.warning(f"Excel file not found, skipping formatting: {excel_path}")
        return

    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        if ws.max_row < 2:
            logger.info("Excel has no data rows → no formatting applied")
            return

        # Freeze header row (row 1)
        ws.freeze_panes = "A2"

        # Auto-filter on entire used range
        if ws.dimensions:
            ws.auto_filter.ref = ws.dimensions

        # Find index of 'confidence_category' column
        conf_col_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == "confidence_category":
                conf_col_idx = col_idx
                break

        if conf_col_idx is None:
            logger.warning("Column 'confidence_category' not found → skipping row coloring")
        else:
            # Apply row colors based on confidence category
            for row_idx in range(2, ws.max_row + 1):
                category_cell = ws.cell(row=row_idx, column=conf_col_idx)
                category = category_cell.value
                if category in FILL_COLORS:
                    fill = FILL_COLORS[category]
                    for cell in ws[row_idx]:
                        cell.fill = fill

        wb.save(excel_path)
        logger.info(f"Excel formatting applied successfully: freeze panes, filter, confidence colors")

    except Exception as e:
        logger.warning(f"Excel formatting failed (file still saved): {type(e).__name__} - {str(e)}")